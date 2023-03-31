import os
import zipfile

import openpyxl
import pytest
from PyPDF2 import PdfReader

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_PATH = os.path.abspath(os.path.join(CURRENT_DIR, "..", "SourceForArchive"))


@pytest.fixture
def files_to_archive():
    current_dir = CURRENT_DIR
    source_path = SOURCE_PATH
    resources = os.path.abspath(os.path.join(current_dir, "..", "resources"))
    with zipfile.ZipFile(f'{resources}/file.zip', 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=5) as newzipfile:
        for root, dirs, files in os.walk(source_path):
            for filename in files:
                newzipfile.write(os.path.abspath(os.path.join(source_path, filename)), filename)
    yield files_to_archive
    os.remove(f'{resources}/file.zip')


def file_content_pdf_in_archive():
    current_dir = CURRENT_DIR
    resources = os.path.abspath(os.path.join(current_dir, "..", "resources"))
    with zipfile.ZipFile(f'{resources}/file.zip') as pdffile:
        pdf_open = pdffile.open('Sample_pdf.pdf')
        pdf_reader = PdfReader(pdf_open)
        pdf_pages = len(pdf_reader.pages)
        pdf_metadata = pdf_reader.metadata
        pdf_open.close()
        return [pdf_pages, pdf_metadata]


def file_content_csv_in_archive():
    current_dir = CURRENT_DIR
    resources = os.path.abspath(os.path.join(current_dir, "..", "resources"))
    with zipfile.ZipFile(f'{resources}/file.zip') as csvfile:
        person = csvfile.open("Sample_csv.csv")
        csv_rows = 0
        persons = []
        for row in person:
            csv_rows += 1
            persons.append(row)
        return [csv_rows, persons]


def file_content_xlsx_in_archive():
    current_dir = CURRENT_DIR
    resources = os.path.abspath(os.path.join(current_dir, "..", "resources"))
    with zipfile.ZipFile(f'{resources}/file.zip') as xlsxfa:
        xlsxfile = xlsxfa.open("Sample_xlsx.xlsx")
        xlsx_reader = openpyxl.load_workbook(xlsxfile)
        sheet = xlsx_reader.active
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                sheet_content.append(value)
        return sheet_content


def file_pdf_content():
    with open(f"{SOURCE_PATH}/sample_pdf.pdf", "rb") as pdffile:
        pdf_reader = PdfReader(pdffile)
        pdf_pages = len(pdf_reader.pages)
        pdf_metadata = pdf_reader.metadata
        return [pdf_pages, pdf_metadata]


def file_csv_content():
    with open(f"{SOURCE_PATH}/Sample_csv.csv", "rb") as csvfile:
        csv_rows = 0
        persons = []
        for row in csvfile:
            csv_rows += 1
            persons.append(row)
        return [csv_rows, persons]


def file_xlsx_content():
    with open(f"{SOURCE_PATH}/Sample_xlsx.xlsx", "rb") as xlsxfile:
        xlsx_reader = openpyxl.load_workbook(xlsxfile)
        sheet = xlsx_reader.active
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                sheet_content.append(value)
        return sheet_content


def test_file_archive(files_to_archive):
    assert file_xlsx_content() == file_content_xlsx_in_archive(), 'Содержимое файла xlsx в папке ресуров и в архиве ' \
                                                                  'отличается'
    assert file_csv_content()[0] == file_content_csv_in_archive()[0], 'Количество строк файла csv в папке ресуров и в ' \
                                                                      'архиве отличается'
    assert file_csv_content()[1] == file_content_csv_in_archive()[1], 'Содержание записей файла csv в папке ресуров и ' \
                                                                      'в архиве отличается'
    assert file_pdf_content()[0] == file_content_pdf_in_archive()[0], 'Количество страниц файла pdf в папке ресуров и ' \
                                                                      'в архиве отличается'
    assert file_pdf_content()[1] == file_content_pdf_in_archive()[1], 'metadata файла pdf в папке ресуров и в архиве ' \
                                                                      'отличается'

